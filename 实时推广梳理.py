# ======================= 影刀环境 - 实时推广梳理（无依赖库版） =======================
# 仅依赖openpyxl（如影刀环境自带），否则请先安装openpyxl
# 不使用pandas、numpy、matplotlib等库，全部用原生Python实现

import os
import sys
from datetime import datetime, timedelta
import re

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("错误：未检测到openpyxl库，请先安装openpyxl\n")
    output1 = "未检测到openpyxl库"
    output2 = output3 = ""
    result = False
    raise SystemExit(1)

# 文件路径配置
file_path = r"E:\文档\佰穗旗舰店推广监控.xlsx"

# 列名映射规则
COLUMN_MAPPING = {
    '时间': ['时间', '日期', 'timestamp', 'date', '日期时间'],
    '店铺名称': ['店铺', '门店', 'shop', 'store', '店铺名称'],
    '计划名称': ['计划名称', '计划名', '计划', 'campaign', '计划类型', '推广计划 ID', '计划ID'],
    '花费': ['花费', '消耗', '支出', 'spend', 'cost', '消耗金额'],
    '日预算': ['日预算', '每日预算', 'daily_budget'],
    '加购成本': ['加购成本', '加购费用', 'add_to_cart_cost'],
    '加购总数': ['加购总数', '加购量', 'add_to_cart_total', '总加购数', '加购件数'],
    '总订单金额': ['总订单金额', '订单总额', 'gmv', 'total_order_value', '支付金额'],
    '投产比': ['投产比', 'roi', '投入产出比'],
    '平均点击成本': ['平均点击成本', '点击成本', 'cpc', 'avg_click_cost']
}

# ======================= 读取Excel数据 =======================
if not os.path.exists(file_path):
    output1 = f"错误：文件不存在 - {file_path}"
    output2 = output3 = ""
    result = False
else:
    try:
        wb = load_workbook(file_path, data_only=True)
        if '快车' not in wb.sheetnames or '智能投放' not in wb.sheetnames:
            output1 = "错误：缺少必要的工作表（快车/智能投放）"
            output2 = output3 = ""
            result = False
            raise SystemExit(1)
        sheet_kuaiche = wb['快车']
        sheet_zhitou = wb['智能投放']
    except Exception as e:
        output1 = f"错误：读取Excel失败 - {str(e)}"
        output2 = output3 = ""
        result = False
        raise SystemExit(1)

    # ======================= 解析表头 =======================
    def get_header_map(sheet):
        header_row = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if any(row):
                header_row = i
                break
        if header_row is None:
            return {}, 0
        headers = [str(cell).strip() if cell is not None else '' for cell in list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]]
        col_map = {}
        for std, names in COLUMN_MAPPING.items():
            for idx, col in enumerate(headers):
                col_clean = re.sub(r'\W+', '', col).lower()
                for n in names + [std]:
                    n_clean = re.sub(r'\W+', '', n).lower()
                    if n_clean == col_clean or n_clean in col_clean:
                        col_map[std] = idx
                        break
                if std in col_map:
                    break
        # 只保留必要映射输出
        # print(f"列名映射({sheet.title}): {col_map}")
        return col_map, header_row

    # ======================= 读取数据为列表 =======================
    def sheet_to_dicts(sheet, col_map, header_row):
        data = []
        for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
            row_dict = {}
            for std, idx in col_map.items():
                if idx < len(row):
                    row_dict[std] = row[idx]
                else:
                    row_dict[std] = None
            data.append(row_dict)
        # print(f"读取sheet({sheet.title})共{len(data)}行，前5行样例: {data[:5]}")
        return data

    col_map_kuaiche, header_row_kuaiche = get_header_map(sheet_kuaiche)
    col_map_zhitou, header_row_zhitou = get_header_map(sheet_zhitou)
    data_kuaiche = sheet_to_dicts(sheet_kuaiche, col_map_kuaiche, header_row_kuaiche)
    data_zhitou = sheet_to_dicts(sheet_zhitou, col_map_zhitou, header_row_zhitou)

    # ======================= 获取所有店铺 =======================
    shops = set()
    for row in data_kuaiche:
        if row.get('店铺名称'):
            shops.add(row['店铺名称'])
    for row in data_zhitou:
        if row.get('店铺名称'):
            shops.add(row['店铺名称'])
    shops = sorted(list(shops))
    # print(f"所有店铺列表: {shops}")

    # ======================= 辅助：时间标准化 =======================
    def parse_time(val):
        if isinstance(val, datetime):
            return val
        if isinstance(val, (int, float)):
            try:
                return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val) - 2)
            except:
                return None
        if isinstance(val, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H", "%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%Y.%m.%d"):
                try:
                    return datetime.strptime(val.strip(), fmt)
                except:
                    continue
        return None

    # ======================= 主体分析 =======================
    shop_reports = {}
    for shop in shops:
        # print(f"\n==== 分析店铺: {shop} ====")
        report = f"# {shop}数据分析报告\n\n"
        report += f"**分析时间：** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        kuaiche_alerts = []
        zhitou_alerts = []

        # 处理快车
        shop_kuaiche = [r for r in data_kuaiche if r.get('店铺名称') == shop]
        for r in shop_kuaiche:
            r['时间'] = parse_time(r.get('时间'))
        shop_kuaiche = [r for r in shop_kuaiche if r['时间']]
        shop_kuaiche.sort(key=lambda x: x['时间'])
        if shop_kuaiche:
            latest_time = shop_kuaiche[-1]['时间']
            for plan in set(r.get('计划名称') for r in shop_kuaiche if r.get('时间') == latest_time):
                plan_rows = [r for r in shop_kuaiche if r.get('计划名称') == plan and r.get('时间') == latest_time]
                if not plan_rows:
                    continue
                purchase_cost = plan_rows[0].get('加购成本')
                try:
                    purchase_cost = float(purchase_cost)
                except:
                    continue
                if purchase_cost > 40:
                    kuaiche_alerts.append(f"快车-{plan}，加购成本为{purchase_cost:.2f}，严重偏高，请抓紧调整")
                elif 0 < purchase_cost < 10:
                    kuaiche_alerts.append(f"快车-{plan}，加购成本为{purchase_cost:.2f}，严重偏低，请抓紧调整")
            if len(set(r['时间'].date() for r in shop_kuaiche)) > 1:
                yesterday = latest_time.date() - timedelta(days=1)
                for plan in set(r.get('计划名称') for r in shop_kuaiche if r.get('时间') == latest_time):
                    today_row = [r for r in shop_kuaiche if r.get('计划名称') == plan and r.get('时间') == latest_time]
                    yest_row = [r for r in shop_kuaiche if r.get('计划名称') == plan and r.get('时间') and r['时间'].date() == yesterday]
                    if today_row and yest_row:
                        try:
                            today_val = float(today_row[0].get('加购总数', 0) or 0)
                            yest_val = float(yest_row[0].get('加购总数', 0) or 0)
                        except:
                            continue
                        if yest_val > 0:
                            rate = (today_val - yest_val) / yest_val
                            if abs(rate) >= 0.2:
                                status = "增长" if rate > 0 else "下滑"
                                percent = abs(rate) * 100
                                hour = latest_time.hour
                                kuaiche_alerts.append(f"快车-{plan}截至{hour}点，加购总额为{today_val}，较同期{yest_val}，{status}{percent:.1f}%")
        # 处理智投
        shop_zhitou = [r for r in data_zhitou if r.get('店铺名称') == shop]
        for r in shop_zhitou:
            r['时间'] = parse_time(r.get('时间'))
        shop_zhitou = [r for r in shop_zhitou if r['时间']]
        shop_zhitou.sort(key=lambda x: x['时间'])
        if shop_zhitou:
            latest_time = shop_zhitou[-1]['时间']
            if len(set(r['时间'].date() for r in shop_zhitou)) > 1:
                yesterday = latest_time.date() - timedelta(days=1)
                for plan in set(r.get('计划名称') for r in shop_zhitou if r.get('时间') == latest_time):
                    today_row = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') == latest_time]
                    yest_row = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') and r['时间'].date() == yesterday]
                    if today_row and yest_row:
                        try:
                            today_val = float(today_row[0].get('花费', 0) or 0)
                            yest_val = float(yest_row[0].get('花费', 0) or 0)
                        except:
                            continue
                        if yest_val > 0:
                            rate = (today_val - yest_val) / yest_val
                            if abs(rate) >= 0.4:
                                status = "增长" if rate > 0 else "下滑"
                                percent = abs(rate) * 100
                                hour = latest_time.hour
                                zhitou_alerts.append(f"智投-{plan}截至{hour}点，花费总额为{today_val}，较同期{yest_val}，{status}{percent:.1f}%")
            if len(set(r['时间'].date() for r in shop_zhitou)) > 1:
                yesterday = latest_time.date() - timedelta(days=1)
                for plan in set(r.get('计划名称') for r in shop_zhitou if r.get('时间') == latest_time):
                    today_row = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') == latest_time]
                    yest_row = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') and r['时间'].date() == yesterday]
                    if today_row and yest_row:
                        try:
                            today_val = float(today_row[0].get('总订单金额', 0) or 0)
                            yest_val = float(yest_row[0].get('总订单金额', 0) or 0)
                        except:
                            continue
                        if yest_val > 0:
                            rate = (today_val - yest_val) / yest_val
                            if abs(rate) >= 0.4:
                                status = "增长" if rate > 0 else "下滑"
                                percent = abs(rate) * 100
                                hour = latest_time.hour
                                zhitou_alerts.append(f"智投-{plan}截至{hour}点，订单金额为{today_val}，较同期{yest_val}，{status}{percent:.1f}%")
            if len(set(r['时间'].date() for r in shop_zhitou)) > 1:
                yesterday = latest_time.date() - timedelta(days=1)
                for plan in set(r.get('计划名称') for r in shop_zhitou if r.get('时间') == latest_time):
                    today_row = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') == latest_time]
                    yest_row = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') and r['时间'].date() == yesterday]
                    if today_row and yest_row:
                        try:
                            today_val = float(today_row[0].get('投产比', 0) or 0)
                            yest_val = float(yest_row[0].get('投产比', 0) or 0)
                        except:
                            continue
                        if yest_val > 0:
                            rate = (today_val - yest_val) / yest_val
                            if abs(rate) >= 0.4:
                                status = "增长" if rate > 0 else "下滑"
                                percent = abs(rate) * 100
                                hour = latest_time.hour
                                zhitou_alerts.append(f"智投-{plan}截至{hour}点，投产比{today_val:.2f}，较同期{yest_val:.2f}，{status}{percent:.1f}%")
            for plan in set(r.get('计划名称') for r in shop_zhitou if r.get('时间') == latest_time):
                plan_rows = [r for r in shop_zhitou if r.get('计划名称') == plan and r.get('时间') == latest_time]
                if not plan_rows:
                    continue
                avg_cost = plan_rows[0].get('平均点击成本')
                try:
                    avg_cost = float(avg_cost)
                except:
                    continue
                if avg_cost > 12:
                    hour = latest_time.hour
                    zhitou_alerts.append(f"智投-{plan}截至{hour}点，平均点击成本为{avg_cost:.2f}远超12，尽快调整")
        # 汇总报告
        if not kuaiche_alerts and not zhitou_alerts:
            report += "**状态：** 正常，无异常情况需要关注。"
        else:
            if kuaiche_alerts:
                report += "## 🚨 快车推广异常\n\n"
                for i, alert in enumerate(kuaiche_alerts, 1):
                    report += f"{i}. {alert}\n\n"
            if zhitou_alerts:
                report += "## 🚨 智能投放异常\n\n"
                for i, alert in enumerate(zhitou_alerts, 1):
                    report += f"{i}. {alert}\n\n"
        shop_reports[shop] = report

    # ======================= 输出变量分配 =======================
    shop_list = sorted(shop_reports.keys())
    output1 = shop_reports[shop_list[0]] if len(shop_list) >= 1 else ""
    output2 = shop_reports[shop_list[1]] if len(shop_list) >= 2 else ""
    output3 = ""
    if len(shop_list) >= 3:
        for shop in shop_list[2:]:
            output3 += shop_reports[shop] + "\n\n" + "="*50 + "\n\n"
        output3 = output3.rstrip("\n\n" + "="*50 + "\n\n")
    result = any(shop_reports.values()) 